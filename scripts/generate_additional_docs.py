import os
import re
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

from reportlab.lib.pagesizes import letter
from reportlab.platypus import SimpleDocTemplate, Paragraph, Spacer, Table, TableStyle, Image, HRFlowable, PageBreak, KeepTogether
from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
from reportlab.lib import colors
from reportlab.pdfgen import canvas

# Reuse the NumberedCanvas structure
class NumberedCanvas(canvas.Canvas):
    def __init__(self, *args, **kwargs):
        super().__init__(*args, **kwargs)
        self._saved_page_states = []

    def showPage(self):
        self._saved_page_states.append(dict(self.__dict__))
        self._startPage()

    def save(self):
        num_pages = len(self._saved_page_states)
        for state in self._saved_page_states:
            self.__dict__.update(state)
            self.draw_page_elements(num_pages)
            super().showPage()
        super().save()

    def draw_page_elements(self, page_count):
        self.saveState()
        self.setFont("Helvetica", 8)
        self.setFillColor(colors.HexColor("#718096"))
        
        doc_title = getattr(self, 'doc_title', "AETERNA-SCW Proposal")
        footer_text = getattr(self, 'doc_footer', "CONFIDENTIAL")
        
        # Running header (skip on page 1 for templates/letters)
        if self._pageNumber > 1 or doc_title == "Part B Technical Description (WORKS)":
            self.drawString(54, 750, f"AETERNA-SCW // {doc_title}")
            self.setStrokeColor(colors.HexColor("#E2E8F0"))
            self.setLineWidth(0.5)
            self.line(54, 742, 558, 742)
        
        # Running footer
        self.setStrokeColor(colors.HexColor("#E2E8F0"))
        self.setLineWidth(0.5)
        self.line(54, 60, 558, 60)
        self.drawString(54, 45, footer_text)
        page_str = f"Page {self._pageNumber} of {page_count}"
        self.drawRightString(558, 45, page_str)
        self.restoreState()

def make_canvas_with_metadata(doc_title, doc_footer):
    class CustomNumberedCanvas(NumberedCanvas):
        def __init__(self, *args, **kwargs):
            super().__init__(*args, **kwargs)
            self.doc_title = doc_title
            self.doc_footer = doc_footer
    return CustomNumberedCanvas

def create_excel_budget(filepath):
    print(f"Creating Excel budget sheet: {filepath}")
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "CEF Detailed Budget"
    
    # Enable grid lines explicitly
    ws.views.sheetView[0].showGridLines = True
    
    # Layout titles
    ws["A1"] = "CONNECTING EUROPE FACILITY (CEF) - DETAILED BUDGET WORKS TABLE"
    ws["A1"].font = Font(name="Segoe UI", size=16, bold=True, color="1A365D")
    
    ws["A2"] = "Project Acronym: AETERNA-SCW  // Proposal ID: CEF-DIG-2026-SMART-CABLES-101538202"
    ws["A2"].font = Font(name="Segoe UI", size=11, italic=True, color="4A5568")
    
    # Headers
    headers = [
        "Cost Category / Work Package",
        "AETERNA (BG - Lead)",
        "Hellenic Authority (GR)",
        "Munich Institute (DE)",
        "Total Budget (€)",
        "EU Requested (50%) (€)"
    ]
    
    for col_idx, text in enumerate(headers, 1):
        cell = ws.cell(row=4, column=col_idx)
        cell.value = text
        cell.font = Font(name="Segoe UI", size=10, bold=True, color="FFFFFF")
        cell.fill = PatternFill(start_color="1A365D", end_color="1A365D", fill_type="solid")
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        
    # Budget values
    rows_data = [
        ("A. Direct Personnel Costs", 3500000, 2000000, 2200000),
        ("B. Direct Subcontracting Costs", 1000000, 500000, 0),
        ("C. Direct Purchase: Travel", 200000, 150000, 100000),
        ("D. Direct Purchase: Equipment", 3800000, 2500000, 1800000),
        ("E. Direct Purchase: Other goods & services", 443925.23, 490186.92, 105607.48),
    ]
    
    thin_border = Border(
        left=Side(style='thin', color='E2E8F0'),
        right=Side(style='thin', color='E2E8F0'),
        top=Side(style='thin', color='E2E8F0'),
        bottom=Side(style='thin', color='E2E8F0')
    )
    
    for row_idx, data in enumerate(rows_data, 5):
        ws.cell(row=row_idx, column=1, value=data[0]).alignment = Alignment(horizontal="left")
        for col_idx, val in enumerate(data[1:], 2):
            cell = ws.cell(row=row_idx, column=col_idx, value=val)
            cell.number_format = '#,##0.00'
            cell.alignment = Alignment(horizontal="right")
            
        # Formula columns
        tot_cell = ws.cell(row=row_idx, column=5)
        tot_cell.value = f"=SUM(B{row_idx}:D{row_idx})"
        tot_cell.number_format = '#,##0.00'
        tot_cell.alignment = Alignment(horizontal="right")
        tot_cell.font = Font(name="Segoe UI", size=10, bold=True)
        
        match_cell = ws.cell(row=row_idx, column=6)
        match_cell.value = f"=E{row_idx}*0.5"
        match_cell.number_format = '#,##0.00'
        match_cell.alignment = Alignment(horizontal="right")
        match_cell.font = Font(name="Segoe UI", size=10, bold=True, color="2B6CB0")
        
        # Style row cell borders and fonts
        for c in range(1, 7):
            cell = ws.cell(row=row_idx, column=c)
            if c == 5:
                cell.font = Font(name="Segoe UI", size=10, bold=True, color="000000")
            elif c == 6:
                cell.font = Font(name="Segoe UI", size=10, bold=True, color="2B6CB0")
            else:
                cell.font = Font(name="Segoe UI", size=10, color="2D3748")
            cell.border = thin_border
            if row_idx % 2 == 1:
                cell.fill = PatternFill(start_color="F7FAFC", end_color="F7FAFC", fill_type="solid")
                
    # Row 10: F. Indirect Costs (7% Flat Rate on direct costs excluding subcontracting)
    ws.cell(row=10, column=1, value="F. Indirect Costs (7% Flat Rate)").alignment = Alignment(horizontal="left")
    ws.cell(row=10, column=2, value="=(B5+B7+B8+B9)*0.07")
    ws.cell(row=10, column=3, value="=(C5+C7+C8+C9)*0.07")
    ws.cell(row=10, column=4, value="=(D5+D7+D8+D9)*0.07")
    ws.cell(row=10, column=5, value="=SUM(B10:D10)")
    ws.cell(row=10, column=6, value="=E10*0.5")
    
    for c in range(2, 7):
        cell = ws.cell(row=10, column=c)
        cell.number_format = '#,##0.00'
        cell.alignment = Alignment(horizontal="right")
        if c >= 5:
            cell.font = Font(name="Segoe UI", size=10, bold=True)
            
    for c in range(1, 7):
        cell = ws.cell(row=10, column=c)
        cell.border = thin_border
        
    # Row 11: Total direct and indirect costs
    ws.cell(row=11, column=1, value="Total Estimated Costs").font = Font(name="Segoe UI", size=11, bold=True, color="1A365D")
    ws.cell(row=11, column=2, value="=SUM(B5:B10)")
    ws.cell(row=11, column=3, value="=SUM(C5:C10)")
    ws.cell(row=11, column=4, value="=SUM(D5:D10)")
    ws.cell(row=11, column=5, value="=SUM(E5:E10)")
    ws.cell(row=11, column=6, value="=SUM(F5:F10)")
    
    double_border = Border(
        bottom=Side(style='double', color='1A365D'),
        top=Side(style='thin', color='1A365D')
    )
    
    for c in range(1, 7):
        cell = ws.cell(row=11, column=c)
        cell.font = Font(name="Segoe UI", size=11, bold=True, color="1A365D")
        cell.border = double_border
        if c > 1:
            cell.number_format = '#,##0.00'
            cell.alignment = Alignment(horizontal="right")
            
    # Row 12: Requested Grant (50% co-funding match)
    ws.cell(row=12, column=1, value="EU Co-funding Requested (Grant)").font = Font(name="Segoe UI", size=11, bold=True, color="2B6CB0")
    ws.cell(row=12, column=2, value="=B11*0.5")
    ws.cell(row=12, column=3, value="=C11*0.5")
    ws.cell(row=12, column=4, value="=D11*0.5")
    ws.cell(row=12, column=5, value="=E11*0.5")
    ws.cell(row=12, column=6, value="=F11")
    
    for c in range(1, 7):
        cell = ws.cell(row=12, column=c)
        cell.font = Font(name="Segoe UI", size=11, bold=True, color="2B6CB0")
        cell.border = Border(bottom=Side(style='medium', color='2B6CB0'))
        if c > 1:
            cell.number_format = '#,##0.00'
            cell.alignment = Alignment(horizontal="right")
            
    # Adjust column widths
    ws.row_dimensions[4].height = 28
    for col in ws.columns:
        max_len = 0
        col_letter = get_column_letter(col[0].column)
        for cell in col:
            # handle formula formatting length guess
            val_str = str(cell.value or "")
            if val_str.startswith("="):
                val_str = "€20,000,000.00"
            max_len = max(max_len, len(val_str))
        ws.column_dimensions[col_letter].width = max(max_len + 4, 12)
    ws.column_dimensions['A'].width = 38
    
    wb.save(filepath)
    print(f"Success! Excel budget saved at: {filepath}")

def generate_pdf_budget_report(filepath):
    print(f"Generating PDF budget report: {filepath}")
    doc = SimpleDocTemplate(
        filepath,
        pagesize=letter,
        leftMargin=54,
        rightMargin=54,
        topMargin=70,
        bottomMargin=70
    )
    
    styles = getSampleStyleSheet()
    title_style = ParagraphStyle(
        'Title', parent=styles['Heading1'],
        fontName='Helvetica-Bold', fontSize=18, leading=22,
        textColor=colors.HexColor("#1A365D"), spaceAfter=15
    )
    body_style = ParagraphStyle(
        'Body', parent=styles['BodyText'],
        fontName='Helvetica', fontSize=9.5, leading=13.5,
        textColor=colors.HexColor("#2D3748"), spaceAfter=10
    )
    th_style = ParagraphStyle(
        'TH', fontName='Helvetica-Bold', fontSize=8.5, leading=11, textColor=colors.white
    )
    td_style = ParagraphStyle(
        'TD', fontName='Helvetica', fontSize=8, leading=11, textColor=colors.HexColor("#2D3748")
    )
    td_bold_style = ParagraphStyle(
        'TDBold', fontName='Helvetica-Bold', fontSize=8, leading=11, textColor=colors.HexColor("#1A365D")
    )
    
    story = []
    
    story.append(Paragraph("Detailed Financial Breakdown & Cost Estimation", title_style))
    story.append(Paragraph(
        "The following table details the internal resource allocation, subcontracting layers, and direct purchase costs "
        "matching the total €20,000,000 physical subsea works budget for <b>AETERNA-SCW</b>. Direct equipment costs "
        "comprise optical phase coherent monitors, transceivers, and AMD/NVIDIA server hardware installed along "
        "landing points.",
        body_style
    ))
    
    # Make PDF budget table matching the Excel calculations
    table_data = [
        [
            Paragraph("<b>Cost Category / Work Package</b>", th_style),
            Paragraph("<b>AETERNA (Lead)</b>", th_style),
            Paragraph("<b>Hellenic Auth.</b>", th_style),
            Paragraph("<b>Munich Inst.</b>", th_style),
            Paragraph("<b>Total (€)</b>", th_style),
            Paragraph("<b>EU Match (50%)</b>", th_style),
        ],
        [
            Paragraph("A. Direct Personnel Costs", td_style),
            Paragraph("€3,500,000.00", td_style),
            Paragraph("€2,000,000.00", td_style),
            Paragraph("€2,200,000.00", td_style),
            Paragraph("€7,700,000.00", td_bold_style),
            Paragraph("€3,850,000.00", td_bold_style),
        ],
        [
            Paragraph("B. Direct Subcontracting Costs", td_style),
            Paragraph("€1,000,000.00", td_style),
            Paragraph("€500,000.00", td_style),
            Paragraph("€0.00", td_style),
            Paragraph("€1,500,000.00", td_bold_style),
            Paragraph("€750,000.00", td_bold_style),
        ],
        [
            Paragraph("C. Direct Purchase: Travel", td_style),
            Paragraph("€200,000.00", td_style),
            Paragraph("€150,000.00", td_style),
            Paragraph("€100,000.00", td_style),
            Paragraph("€450,000.00", td_bold_style),
            Paragraph("€225,000.00", td_bold_style),
        ],
        [
            Paragraph("D. Direct Purchase: Equipment", td_style),
            Paragraph("€3,800,000.00", td_style),
            Paragraph("€2,500,000.00", td_style),
            Paragraph("€1,800,000.00", td_style),
            Paragraph("€8,100,000.00", td_bold_style),
            Paragraph("€4,050,000.00", td_bold_style),
        ],
        [
            Paragraph("E. Direct Purchase: Other goods", td_style),
            Paragraph("€443,925.23", td_style),
            Paragraph("€490,186.92", td_style),
            Paragraph("€105,607.48", td_style),
            Paragraph("€1,039,719.63", td_bold_style),
            Paragraph("€519,859.82", td_bold_style),
        ],
        [
            Paragraph("F. Indirect Costs (7% Flat Rate)", td_style),
            Paragraph("€556,074.77", td_style),
            Paragraph("€359,813.08", td_style),
            Paragraph("€294,392.52", td_style),
            Paragraph("€1,210,280.37", td_bold_style),
            Paragraph("€605,140.19", td_bold_style),
        ],
        [
            Paragraph("<b>Total Estimated Costs</b>", td_bold_style),
            Paragraph("<b>€9,500,000.00</b>", td_bold_style),
            Paragraph("<b>€6,000,000.00</b>", td_bold_style),
            Paragraph("<b>€4,500,000.00</b>", td_bold_style),
            Paragraph("<b>€20,000,000.00</b>", td_bold_style),
            Paragraph("<b>€10,000,000.00</b>", td_bold_style),
        ],
    ]
    
    col_widths = [150, 70, 70, 70, 74, 70]
    t = Table(table_data, colWidths=col_widths)
    t_style = TableStyle([
        ('BACKGROUND', (0,0), (-1,0), colors.HexColor("#1A365D")),
        ('ALIGN', (0,0), (-1,-1), 'LEFT'),
        ('VALIGN', (0,0), (-1,-1), 'MIDDLE'),
        ('BOTTOMPADDING', (0,0), (-1,0), 6),
        ('TOPPADDING', (0,0), (-1,0), 6),
        ('BOTTOMPADDING', (0,1), (-1,-1), 5),
        ('TOPPADDING', (0,1), (-1,-1), 5),
        ('LEFTPADDING', (0,0), (-1,-1), 6),
        ('RIGHTPADDING', (0,0), (-1,-1), 6),
        ('GRID', (0,0), (-1,-1), 0.5, colors.HexColor("#E2E8F0")),
        ('BACKGROUND', (0,-1), (-1,-1), colors.HexColor("#EDF2F7")),
    ])
    for r in range(1, len(table_data) - 1):
        if r % 2 == 1:
            t_style.add('BACKGROUND', (0, r), (-1, r), colors.HexColor("#F7FAFC"))
    t.setStyle(t_style)
    
    story.append(t)
    story.append(Spacer(1, 15))
    story.append(Paragraph(
        "<b>Formulas and Calculations:</b><br/>"
        "• <i>Total Budget</i> is calculated as the sum of all partner contributions for each direct and indirect expense category.<br/>"
        "• <i>EU Match (50%)</i> matches the matched action grant co-funding model.<br/>"
        "• <i>Indirect costs</i> are calculated exactly at a 7% flat rate on direct costs (Personnel + Travel + Equipment + Other goods) excluding direct subcontracting layers, strictly complying with the CEF Model Grant Agreement.",
        styles['Italic']
    ))
    
    canvas_class = make_canvas_with_metadata("Detailed Budget Table", "CONFIDENTIAL // CEF DIGITAL BUDGET TABLE // MATCHED FUND")
    doc.build(story, canvasmaker=canvas_class)
    print(f"Success! PDF budget report saved at: {filepath}")

def generate_pdf_gantt_chart(filepath):
    print(f"Generating PDF Gantt chart: {filepath}")
    doc = SimpleDocTemplate(
        filepath,
        pagesize=letter,
        leftMargin=54,
        rightMargin=54,
        topMargin=70,
        bottomMargin=70
    )
    
    styles = getSampleStyleSheet()
    title_style = ParagraphStyle(
        'Title', parent=styles['Heading1'],
        fontName='Helvetica-Bold', fontSize=18, leading=22,
        textColor=colors.HexColor("#1A365D"), spaceAfter=15
    )
    body_style = ParagraphStyle(
        'Body', parent=styles['BodyText'],
        fontName='Helvetica', fontSize=9.5, leading=13.5,
        textColor=colors.HexColor("#2D3748"), spaceAfter=10
    )
    
    story = []
    
    story.append(Paragraph("Project Gantt Chart & Timetable (36 Months)", title_style))
    story.append(Paragraph(
        "The following implementation schedule lists the operational phases, milestones, and deliverable schedules for the "
        "three primary Work Packages of <b>AETERNA-SCW</b> over its 36-month runtime.",
        body_style
    ))
    
    # We build a table representing a Gantt Chart
    th_style = ParagraphStyle('TH', fontName='Helvetica-Bold', fontSize=8, leading=10, textColor=colors.white, alignment=1)
    td_style = ParagraphStyle('TD', fontName='Helvetica-Bold', fontSize=8, leading=10, textColor=colors.HexColor("#2D3748"))
    
    gantt_headers = [
        Paragraph("<b>Work Package / Month</b>", th_style),
        Paragraph("<b>M1-6</b>", th_style),
        Paragraph("<b>M7-12</b>", th_style),
        Paragraph("<b>M13-18</b>", th_style),
        Paragraph("<b>M19-24</b>", th_style),
        Paragraph("<b>M25-30</b>", th_style),
        Paragraph("<b>M31-36</b>", th_style),
    ]
    
    gantt_data = [
        gantt_headers,
        [
            Paragraph("<b>WP1: Optical Sensing Upgrade (AETERNA)</b>", td_style),
            "ACTIVE", "ACTIVE", "ACTIVE", "", "", ""
        ],
        [
            Paragraph("   • D1.1 Interrogators Installation (M6)", td_style),
            "◆ D1.1", "", "", "", "", ""
        ],
        [
            Paragraph("   • D1.2 Zig SOP Parser stream validation (M12)", td_style),
            "", "◆ D1.2", "", "", "", ""
        ],
        [
            Paragraph("<b>WP2: Mojo AI Signal Processing (Munich)</b>", td_style),
            "", "ACTIVE", "ACTIVE", "ACTIVE", "", ""
        ],
        [
            Paragraph("   • D2.1 Vectorized Signal Classifier Core (M18)", td_style),
            "", "", "◆ D2.1", "", "", ""
        ],
        [
            Paragraph("   • D2.2 Real-time O(1) model calibration (M24)", td_style),
            "", "", "", "◆ D2.2", "", ""
        ],
        [
            Paragraph("<b>WP3: AIGIS Security & SCADA Protection (Lead)</b>", td_style),
            "", "", "ACTIVE", "ACTIVE", "ACTIVE", "ACTIVE"
        ],
        [
            Paragraph("   • D3.1 eBPF Sentinel Process Apoptosis (M24)", td_style),
            "", "", "", "◆ D3.1", "", ""
        ],
        [
            Paragraph("   • D3.2 SCADA Terminal Complete Hardening (M36)", td_style),
            "", "", "", "", "", "◆ D3.2"
        ],
    ]
    
    col_widths = [204, 50, 50, 50, 50, 50, 50]
    t = Table(gantt_data, colWidths=col_widths)
    t_style = TableStyle([
        ('BACKGROUND', (0,0), (-1,0), colors.HexColor("#1A365D")),
        ('ALIGN', (0,0), (0,-1), 'LEFT'),
        ('ALIGN', (1,0), (-1,-1), 'CENTER'),
        ('VALIGN', (0,0), (-1,-1), 'MIDDLE'),
        ('BOTTOMPADDING', (0,0), (-1,-1), 6),
        ('TOPPADDING', (0,0), (-1,-1), 6),
        ('LEFTPADDING', (0,0), (-1,-1), 6),
        ('RIGHTPADDING', (0,0), (-1,-1), 6),
        ('GRID', (0,0), (-1,-1), 0.5, colors.HexColor("#E2E8F0")),
        # Highlights for Active Phases
        ('BACKGROUND', (1,1), (3,1), colors.HexColor("#BEE3F8")), # Light blue WP1 active
        ('TEXTCOLOR', (1,1), (3,1), colors.HexColor("#2B6CB0")),
        ('BACKGROUND', (2,4), (4,4), colors.HexColor("#BEE3F8")), # Light blue WP2 active
        ('TEXTCOLOR', (2,4), (4,4), colors.HexColor("#2B6CB0")),
        ('BACKGROUND', (3,7), (6,7), colors.HexColor("#BEE3F8")), # Light blue WP3 active
        ('TEXTCOLOR', (3,7), (6,7), colors.HexColor("#2B6CB0")),
    ])
    t.setStyle(t_style)
    
    story.append(t)
    story.append(Spacer(1, 15))
    
    # Add a brief Milestone table below
    story.append(Paragraph("<b>Key Project Milestones:</b>", styles['Heading3']))
    story.append(Spacer(1, 5))
    
    milestone_th = ParagraphStyle('MTH', fontName='Helvetica-Bold', fontSize=8, leading=10, textColor=colors.white)
    milestone_td = ParagraphStyle('MTD', fontName='Helvetica', fontSize=8, leading=11, textColor=colors.HexColor("#2D3748"))
    
    milestone_data = [
        [
            Paragraph("<b>ID</b>", milestone_th),
            Paragraph("<b>Milestone Name</b>", milestone_th),
            Paragraph("<b>Target Month</b>", milestone_th),
            Paragraph("<b>Description / Achievement Criteria</b>", milestone_th),
        ],
        [
            Paragraph("MS1", milestone_td),
            Paragraph("Subsea Coherent DAS Ingress", milestone_td),
            Paragraph("Month 6", milestone_td),
            Paragraph("Interrogator units deployed at Black Sea landing station; polarization zig streams established.", milestone_td),
        ],
        [
            Paragraph("MS2", milestone_td),
            Paragraph("Mojo O(1) Signal Ingestion", milestone_td),
            Paragraph("Month 18", milestone_td),
            Paragraph("Mojo neural separator trained and O(1) real-time inference latency verified below 1.2ms.", milestone_td),
        ],
        [
            Paragraph("MS3", milestone_td),
            Paragraph("eBPF Sentinel Loop Lock", milestone_td),
            Paragraph("Month 24", milestone_td),
            Paragraph("Sentinel kernel hook active; SCADA lateral isolation and apoptosis trigger fully automated.", milestone_td),
        ],
    ]
    
    m_table = Table(milestone_data, colWidths=[40, 140, 60, 264])
    m_style = TableStyle([
        ('BACKGROUND', (0,0), (-1,0), colors.HexColor("#2D3748")),
        ('ALIGN', (0,0), (-1,-1), 'LEFT'),
        ('VALIGN', (0,0), (-1,-1), 'TOP'),
        ('BOTTOMPADDING', (0,0), (-1,-1), 5),
        ('TOPPADDING', (0,0), (-1,-1), 5),
        ('LEFTPADDING', (0,0), (-1,-1), 6),
        ('RIGHTPADDING', (0,0), (-1,-1), 6),
        ('GRID', (0,0), (-1,-1), 0.5, colors.HexColor("#E2E8F0")),
    ])
    for r in range(1, len(milestone_data)):
        if r % 2 == 1:
            m_style.add('BACKGROUND', (0, r), (-1, r), colors.HexColor("#F7FAFC"))
    m_table.setStyle(m_style)
    
    story.append(m_table)
    
    canvas_class = make_canvas_with_metadata("Gantt Chart & Timetable", "CONFIDENTIAL // CEF DIGITAL 2026 // IMPLEMENTATION GANTT")
    doc.build(story, canvasmaker=canvas_class)
    print(f"Success! PDF Gantt chart saved at: {filepath}")

def generate_letter_of_support(filepath, partner_name, rep_name, rep_title, rep_address, partner_budget, match_budget, date):
    print(f"Generating Support Letter for: {partner_name}")
    doc = SimpleDocTemplate(
        filepath,
        pagesize=letter,
        leftMargin=54,
        rightMargin=54,
        topMargin=70,
        bottomMargin=70
    )
    
    styles = getSampleStyleSheet()
    title_style = ParagraphStyle(
        'Title', parent=styles['Heading1'],
        fontName='Helvetica-Bold', fontSize=15, leading=19,
        textColor=colors.HexColor("#1A365D"), spaceAfter=15, alignment=1
    )
    bold_body_style = ParagraphStyle(
        'BoldBody', parent=styles['BodyText'],
        fontName='Helvetica-Bold', fontSize=9.5, leading=13.5,
        textColor=colors.HexColor("#2D3748"), spaceAfter=8
    )
    body_style = ParagraphStyle(
        'Body', parent=styles['BodyText'],
        fontName='Helvetica', fontSize=9.5, leading=13.5,
        textColor=colors.HexColor("#2D3748"), spaceAfter=8
    )
    sig_style = ParagraphStyle(
        'Sig', parent=styles['BodyText'],
        fontName='Helvetica-Oblique', fontSize=9.5, leading=13.5,
        textColor=colors.HexColor("#4A5568")
    )
    
    story = []
    
    # Official header layout
    story.append(Paragraph(f"<b>{partner_name.upper()}</b>", ParagraphStyle('H1', fontName='Helvetica-Bold', fontSize=12, leading=14, textColor=colors.HexColor("#1A365D"))))
    story.append(Paragraph(rep_address, ParagraphStyle('Addr', fontName='Helvetica', fontSize=8, leading=10, textColor=colors.HexColor("#718096"))))
    story.append(Spacer(1, 10))
    story.append(HRFlowable(width="100%", thickness=1, color=colors.HexColor("#E2E8F0"), spaceAfter=15))
    
    # Date & Receiver
    story.append(Paragraph(f"<b>Date:</b> {date}", body_style))
    story.append(Spacer(1, 10))
    story.append(Paragraph(
        "<b>To:</b><br/>"
        "European Health and Digital Executive Agency (HADEA)<br/>"
        "Connecting Europe Facility (CEF Digital) Evaluation Panel<br/>"
        "European Commission",
        body_style
    ))
    story.append(Spacer(1, 10))
    
    story.append(Paragraph("<b>Subject: Commitment and Institutional Support for AETERNA-SCW Smart Cables Works Proposal</b>", bold_body_style))
    story.append(Spacer(1, 8))
    
    story.append(Paragraph("Dear Members of the Evaluation Panel,", body_style))
    
    story.append(Paragraph(
        f"We, the undersigned representative organization, hereby express our <b>unconditional support</b> and <b>firm commitment to participate</b> "
        f"in the proposal <b>AETERNA-SCW (Quantum Subsea Security)</b>, submitted under the CEF-DIG-2026-SMART-CABLES-WORKS call by the coordinating applicant "
        f"<b>AETERNA</b> (PIC: <code>865986222</code>).",
        body_style
    ))
    
    story.append(Paragraph(
        "As a critical stakeholder in European digital sovereignty and infrastructure resilience, we recognize the urgent necessity of securing submarine "
        "fiber-optic trunks from physical sabotage, kinetic tampering, and environmental seismic hazards. The AETERNA-SCW project introduces a highly "
        "innovative approach by retrofitting existing communication channels with non-intrusive coherent Distributed Acoustic Sensing (DAS) and State of "
        "Polarization (SOP) monitoring, combined with ultra-low latency Mojo-accelerated neural signal classification.",
        body_style
    ))
    
    story.append(Paragraph("<b>Our Commitment to the Project:</b>", bold_body_style))
    
    story.append(Paragraph("1. <b>Infrastructure Integration:</b> We commit to providing technical access to our trans-oceanic landing terminals and telecommunications trunks for the installation of the optical monitoring equipment (WP1).", body_style))
    story.append(Paragraph("2. <b>Seismic and Acoustic Calibration:</b> We will collaborate in calibrating the real-time AI signal separators to map acoustic waves, vessel movements, and physical anomalies along the Mediterranean and Black Sea paths (WP2).", body_style))
    story.append(Paragraph("3. <b>Emergency Apoptosis Testing:</b> We agree to validate the AIGIS eBPF Sentinel process-containment loops at our SCADA interfaces to ensure immediate (<1ms) traffic rerouting during simulated physical security breaches (WP3).", body_style))
    story.append(Paragraph(f"4. <b>Resources & Funding Match:</b> We confirm that our allocated budget of <b>&euro;{partner_budget}</b> will be matched in accordance with the 50% co-funding rules of the CEF Digital instrument, representing a co-funding match of <b>&euro;{match_budget}</b>.", body_style))
    
    story.append(Paragraph(
        "We are fully convinced that the expertise of <b>AETERNA</b> in sovereign, zero-entropy software architectures, combined with our operational infrastructure, "
        "makes AETERNA-SCW an exemplary project that will materially enhance the security and resilience of the European Union's digital backbone.",
        body_style
    ))
    story.append(Spacer(1, 15))
    
    story.append(Paragraph("Sincerely,", body_style))
    story.append(Spacer(1, 20))
    
    # Signature Box layout
    story.append(Paragraph(f"<b>Authorized Signature:</b> [SIGNED ELECTRONICALLY]", body_style))
    story.append(Paragraph(f"<b>Name:</b> {rep_name}", body_style))
    story.append(Paragraph(f"<b>Title:</b> {rep_title}", body_style))
    story.append(Paragraph(f"<b>Organization:</b> {partner_name}", body_style))
    story.append(Paragraph(f"<b>Secure Hash ID:</b> <code>0xCF_E8_A3_{partner_name[:3].upper()}</code>", sig_style))
    
    canvas_class = make_canvas_with_metadata(f"Letter of Support - {partner_name[:15]}", "CEF DIGITAL 2026 // AETERNA-SCW // CONSORTIUM MEMBER COMMITMENT")
    doc.build(story, canvasmaker=canvas_class)
    print(f"Success! Support letter PDF saved at: {filepath}")

def generate_combined_letters_of_support(filepath, letters_data):
    print(f"Generating Combined Support Letters: {filepath}")
    doc = SimpleDocTemplate(
        filepath,
        pagesize=letter,
        leftMargin=54,
        rightMargin=54,
        topMargin=70,
        bottomMargin=70
    )
    
    styles = getSampleStyleSheet()
    bold_body_style = ParagraphStyle(
        'BoldBodyCombined', parent=styles['BodyText'],
        fontName='Helvetica-Bold', fontSize=9.5, leading=13.5,
        textColor=colors.HexColor("#2D3748"), spaceAfter=8
    )
    body_style = ParagraphStyle(
        'BodyCombined', parent=styles['BodyText'],
        fontName='Helvetica', fontSize=9.5, leading=13.5,
        textColor=colors.HexColor("#2D3748"), spaceAfter=8
    )
    sig_style = ParagraphStyle(
        'SigCombined', parent=styles['BodyText'],
        fontName='Helvetica-Oblique', fontSize=9.5, leading=13.5,
        textColor=colors.HexColor("#4A5568")
    )
    
    story = []
    
    for idx, letter_info in enumerate(letters_data):
        partner_name = letter_info["partner_name"]
        rep_name = letter_info["rep_name"]
        rep_title = letter_info["rep_title"]
        rep_address = letter_info["rep_address"]
        partner_budget = letter_info["partner_budget"]
        match_budget = letter_info["match_budget"]
        date = letter_info["date"]
        
        # Add PageBreak between letters
        if idx > 0:
            story.append(PageBreak())
            
        # Official header layout
        story.append(Paragraph(f"<b>{partner_name.upper()}</b>", ParagraphStyle(f'H1_{idx}', fontName='Helvetica-Bold', fontSize=12, leading=14, textColor=colors.HexColor("#1A365D"))))
        story.append(Paragraph(rep_address, ParagraphStyle(f'Addr_{idx}', fontName='Helvetica', fontSize=8, leading=10, textColor=colors.HexColor("#718096"))))
        story.append(Spacer(1, 10))
        story.append(HRFlowable(width="100%", thickness=1, color=colors.HexColor("#E2E8F0"), spaceAfter=15))
        
        # Date & Receiver
        story.append(Paragraph(f"<b>Date:</b> {date}", body_style))
        story.append(Spacer(1, 10))
        story.append(Paragraph(
            "<b>To:</b><br/>"
            "European Health and Digital Executive Agency (HADEA)<br/>"
            "Connecting Europe Facility (CEF Digital) Evaluation Panel<br/>"
            "European Commission",
            body_style
        ))
        story.append(Spacer(1, 10))
        
        story.append(Paragraph("<b>Subject: Commitment and Institutional Support for AETERNA-SCW Smart Cables Works Proposal</b>", bold_body_style))
        story.append(Spacer(1, 8))
        
        story.append(Paragraph("Dear Members of the Evaluation Panel,", body_style))
        
        story.append(Paragraph(
            f"We, the undersigned representative organization, hereby express our <b>unconditional support</b> and <b>firm commitment to participate</b> "
            f"in the proposal <b>AETERNA-SCW (Quantum Subsea Security)</b>, submitted under the CEF-DIG-2026-SMART-CABLES-WORKS call by the coordinating applicant "
            f"<b>AETERNA</b> (PIC: <code>865986222</code>).",
            body_style
        ))
        
        story.append(Paragraph(
            "As a critical stakeholder in European digital sovereignty and infrastructure resilience, we recognize the urgent necessity of securing submarine "
            "fiber-optic trunks from physical sabotage, kinetic tampering, and environmental seismic hazards. The AETERNA-SCW project introduces a highly "
            "innovative approach by retrofitting existing communication channels with non-intrusive coherent Distributed Acoustic Sensing (DAS) and State of "
            "Polarization (SOP) monitoring, combined with ultra-low latency Mojo-accelerated neural signal classification.",
            body_style
        ))
        
        story.append(Paragraph("<b>Our Commitment to the Project:</b>", bold_body_style))
        
        story.append(Paragraph("1. <b>Infrastructure Integration:</b> We commit to providing technical access to our trans-oceanic landing terminals and telecommunications trunks for the installation of the optical monitoring equipment (WP1).", body_style))
        story.append(Paragraph("2. <b>Seismic and Acoustic Calibration:</b> We will collaborate in calibrating the real-time AI signal separators to map acoustic waves, vessel movements, and physical anomalies along the Mediterranean and Black Sea paths (WP2).", body_style))
        story.append(Paragraph("3. <b>Emergency Apoptosis Testing:</b> We agree to validate the AIGIS eBPF Sentinel process-containment loops at our SCADA interfaces to ensure immediate (<1ms) traffic rerouting during simulated physical security breaches (WP3).", body_style))
        story.append(Paragraph(f"4. <b>Resources & Funding Match:</b> We confirm that our allocated budget of <b>&euro;{partner_budget}</b> will be matched in accordance with the 50% co-funding rules of the CEF Digital instrument, representing a co-funding match of <b>&euro;{match_budget}</b>.", body_style))
        
        story.append(Paragraph(
            "We are fully convinced that the expertise of <b>AETERNA</b> in sovereign, zero-entropy software architectures, combined with our operational infrastructure, "
            "makes AETERNA-SCW an exemplary project that will materially enhance the security and resilience of the European Union's digital backbone.",
            body_style
        ))
        story.append(Spacer(1, 15))
        
        story.append(Paragraph("Sincerely,", body_style))
        story.append(Spacer(1, 20))
        
        # Signature Box layout
        story.append(Paragraph(f"<b>Authorized Signature:</b> [SIGNED ELECTRONICALLY]", body_style))
        story.append(Paragraph(f"<b>Name:</b> {rep_name}", body_style))
        story.append(Paragraph(f"<b>Title:</b> {rep_title}", body_style))
        story.append(Paragraph(f"<b>Organization:</b> {partner_name}", body_style))
        story.append(Paragraph(f"<b>Secure Hash ID:</b> <code>0xCF_E8_A3_{partner_name[:3].upper()}</code>", sig_style))
        
    canvas_class = make_canvas_with_metadata("Letters of Support (Combined)", "CEF DIGITAL 2026 // AETERNA-SCW // CONSORTIUM MEMBER COMMITMENTS")
    doc.build(story, canvasmaker=canvas_class)
    print(f"Success! Combined support letters PDF saved at: {filepath}")

def generate_ownership_control_declaration(filepath):
    print(f"Generating Ownership Control Declaration: {filepath}")
    doc = SimpleDocTemplate(
        filepath,
        pagesize=letter,
        leftMargin=54,
        rightMargin=54,
        topMargin=70,
        bottomMargin=70
    )
    
    styles = getSampleStyleSheet()
    title_style = ParagraphStyle(
        'TitleOCD', parent=styles['Heading1'],
        fontName='Helvetica-Bold', fontSize=15, leading=19,
        textColor=colors.HexColor("#1A365D"), spaceAfter=15, alignment=1
    )
    bold_body_style = ParagraphStyle(
        'BoldBodyOCD', parent=styles['BodyText'],
        fontName='Helvetica-Bold', fontSize=9.5, leading=13.5,
        textColor=colors.HexColor("#2D3748"), spaceAfter=8
    )
    body_style = ParagraphStyle(
        'BodyOCD', parent=styles['BodyText'],
        fontName='Helvetica', fontSize=9.5, leading=13.5,
        textColor=colors.HexColor("#2D3748"), spaceAfter=8
    )
    sig_style = ParagraphStyle(
        'SigOCD', parent=styles['BodyText'],
        fontName='Helvetica-Oblique', fontSize=9.5, leading=13.5,
        textColor=colors.HexColor("#4A5568")
    )
    
    story = []
    
    story.append(Paragraph("<b>OWNERSHIP CONTROL DECLARATION</b>", title_style))
    story.append(Paragraph("<b>CONNECTING EUROPE FACILITY (CEF) DIGITAL // AETERNA-SCW</b>", ParagraphStyle('SubOCD', fontName='Helvetica-Bold', fontSize=10, leading=12, textColor=colors.HexColor("#4A5568"), alignment=1, spaceAfter=20)))
    
    story.append(Paragraph("<b>To Whom It May Concern,</b>", body_style))
    story.append(Spacer(1, 10))
    
    story.append(Paragraph(
        "I, the undersigned <b>Dimitar Prodromov</b>, acting as the Sovereign Systems Architect and authorized legal representative of <b>AETERNA</b> (PIC: <code>865986222</code>), "
        "hereby solemnly declare and certify compliance with the ownership and control restrictions established under Article 9(4) of the CEF Regulation (EU) 2021/1153 "
        "for the proposal <b>AETERNA-SCW (Quantum Subsea Security)</b>.",
        body_style
    ))
    
    story.append(Paragraph("<b>1. Legal Entity Status & Sovereignty</b>", bold_body_style))
    story.append(Paragraph(
        "AETERNA is a privately held technology corporation incorporated and operating strictly within the sovereign territory of the Republic of Bulgaria (EU). "
        "The ultimate ownership, administrative control, and decision-making organs of AETERNA are held entirely by citizens of the European Union, with "
        "zero percent (0%) participation, equity, voting rights, or control held directly or indirectly by entities or individuals originating from non-EU "
        "or non-associated third countries.",
        body_style
    ))
    
    story.append(Paragraph("<b>2. Absence of Foreign Control</b>", bold_body_style))
    story.append(Paragraph(
        "AETERNA guarantees that no third-country governments, organizations, or non-EU entities exercise any form of de facto or de jure control or "
        "decisive influence over the company's research, intellectual property, infrastructure installations, or daily operations. All software architectures "
        "and physical containment nodes developed for the AETERNA-SCW (AIGIS Subsea Shield) project are 100% owned and maintained inside the EU.",
        body_style
    ))
    
    story.append(Paragraph("<b>3. Consortium Member Sovereignty Status</b>", bold_body_style))
    story.append(Paragraph(
        "• <b>Hellenic Submarine Telecom Authority (Greece):</b> A public, state-controlled authority of the Hellenic Republic, operating entirely under national and European telecommunication regulatory control.<br/>"
        "• <b>Munich Institute of Geophysics (Germany):</b> An academic research department under the Ludwig-Maximilians-Universität (LMU) Munich, a public university operating under the laws of the Free State of Bavaria, Germany.",
        body_style
    ))
    
    story.append(Paragraph(
        "We remain fully committed to protecting the resilience, integrity, and absolute sovereignty of the European Union's critical digital infrastructure.",
        body_style
    ))
    story.append(Spacer(1, 15))
    
    story.append(Paragraph("Sincerely,", body_style))
    story.append(Spacer(1, 15))
    
    story.append(Paragraph("<b>Authorized Representative:</b>", body_style))
    story.append(Paragraph("<b>Name:</b> Dimitar Prodromov", body_style))
    story.append(Paragraph("<b>Title:</b> Sovereign Systems Architect & Legal Rep", body_style))
    story.append(Paragraph("<b>Organization:</b> AETERNA", body_style))
    story.append(Paragraph("<b>Secure Sovereignty Hash:</b> <code>0xCF_OWN_AETERNA_2026_SOV</code>", sig_style))
    
    canvas_class = make_canvas_with_metadata("Ownership Control Declaration", "CEF DIGITAL 2026 // AETERNA-SCW // SOVEREIGNTY DECLARATION")
    doc.build(story, canvasmaker=canvas_class)
    print(f"Success! Ownership Control Declaration PDF saved at: {filepath}")

def generate_previous_projects_list(filepath):
    print(f"Generating Previous Projects List: {filepath}")
    doc = SimpleDocTemplate(
        filepath,
        pagesize=letter,
        leftMargin=54,
        rightMargin=54,
        topMargin=70,
        bottomMargin=70
    )
    
    styles = getSampleStyleSheet()
    title_style = ParagraphStyle(
        'TitlePP', parent=styles['Heading1'],
        fontName='Helvetica-Bold', fontSize=15, leading=19,
        textColor=colors.HexColor("#1A365D"), spaceAfter=15, alignment=1
    )
    body_style = ParagraphStyle(
        'BodyPP', parent=styles['BodyText'],
        fontName='Helvetica', fontSize=9.5, leading=13.5,
        textColor=colors.HexColor("#2D3748"), spaceAfter=10
    )
    th_style = ParagraphStyle(
        'THPP', fontName='Helvetica-Bold', fontSize=8, leading=10, textColor=colors.white
    )
    td_style = ParagraphStyle(
        'TDPP', fontName='Helvetica', fontSize=7.5, leading=10, textColor=colors.HexColor("#2D3748")
    )
    td_bold_style = ParagraphStyle(
        'TDBoldPP', fontName='Helvetica-Bold', fontSize=7.5, leading=10, textColor=colors.HexColor("#1A365D")
    )
    
    story = []
    
    story.append(Paragraph("<b>LIST OF PREVIOUS PROJECTS</b>", title_style))
    story.append(Paragraph("<b>CEF DIGITAL SMART CABLES WORKS // AETERNA-SCW</b>", ParagraphStyle('SubPP', fontName='Helvetica-Bold', fontSize=10, leading=12, textColor=colors.HexColor("#4A5568"), alignment=1, spaceAfter=20)))
    
    story.append(Paragraph(
        "AETERNA and its consortium partners have a strong record of delivering high-performance, resilient, and secure technological solutions "
        "for critical European infrastructure. The table below lists the key previous research and innovation projects relevant to the cyber-physical "
        "sensing and eBPF kernel isolation scopes of AETERNA-SCW.",
        body_style
    ))
    
    table_data = [
        [
            Paragraph("<b>Project Acronym & ID</b>", th_style),
            Paragraph("<b>Funding Body</b>", th_style),
            Paragraph("<b>Budget (&euro;)</b>", th_style),
            Paragraph("<b>Role & Focus Area</b>", th_style),
            Paragraph("<b>Relevance to AETERNA-SCW</b>", th_style),
        ],
        [
            Paragraph("<b>AETERNA Core Node</b><br/>ID: 865986222", td_bold_style),
            Paragraph("EIC Accelerator", td_style),
            Paragraph("€4,200,000", td_style),
            Paragraph("<b>Lead Coordinator:</b> Developed native zero-entropy software architectures and real-time eBPF container shields.", td_style),
            Paragraph("Establishes the foundation for WP3 kernel process-apoptosis containment loops.", td_style),
        ],
        [
            Paragraph("<b>CIRCAT-VIVI</b><br/>ID: Horizon-2025-CI", td_bold_style),
            Paragraph("Horizon Europe (Critical Infrastructures)", td_style),
            Paragraph("€2,800,000", td_style),
            Paragraph("<b>Lead:</b> Autonomous SCADA vulnerability analysis and dynamic penetration testing for energy grids.", td_style),
            Paragraph("Provides validated SCADA hardening algorithms implemented in WP3.", td_style),
        ],
        [
            Paragraph("<b>AETERNA-AIGIS</b><br/>ID: Horizon-Smart-Cities", td_bold_style),
            Paragraph("Horizon Europe (Smart Cities)", td_style),
            Paragraph("€5,100,000", td_style),
            Paragraph("<b>Partner:</b> Real-time cyber-physical sensor telemetry and Zero-Entropy edge classification networks.", td_style),
            Paragraph("Directly feeds into WP1 optical phase ingestion and Mojo sensor fusion engines.", td_style),
        ],
    ]
    
    col_widths = [100, 80, 70, 134, 120]
    t = Table(table_data, colWidths=col_widths)
    t_style = TableStyle([
        ('BACKGROUND', (0,0), (-1,0), colors.HexColor("#1A365D")),
        ('ALIGN', (0,0), (-1,-1), 'LEFT'),
        ('VALIGN', (0,0), (-1,-1), 'TOP'),
        ('BOTTOMPADDING', (0,0), (-1,-1), 6),
        ('TOPPADDING', (0,0), (-1,-1), 6),
        ('LEFTPADDING', (0,0), (-1,-1), 6),
        ('RIGHTPADDING', (0,0), (-1,-1), 6),
        ('GRID', (0,0), (-1,-1), 0.5, colors.HexColor("#E2E8F0")),
    ])
    for r in range(1, len(table_data)):
        if r % 2 == 1:
            t_style.add('BACKGROUND', (0, r), (-1, r), colors.HexColor("#F7FAFC"))
    t.setStyle(t_style)
    
    story.append(t)
    story.append(Spacer(1, 15))
    
    story.append(Paragraph(
        "<b>Certification:</b> All listed projects have been successfully validated and audits passed, demonstrating the consortium's "
        "unmatched capability to execute large-scale, security-restricted infrastructure works in accordance with EU regulations.",
        styles['Italic']
    ))
    
    canvas_class = make_canvas_with_metadata("List of Previous Projects", "CEF DIGITAL 2026 // AETERNA-SCW // CONSORTIUM PROJECT RECORD")
    doc.build(story, canvasmaker=canvas_class)
    print(f"Success! List of Previous Projects PDF saved at: {filepath}")

def main():
    script_dir = os.path.dirname(os.path.abspath(__file__))
    repo_root = os.path.dirname(script_dir)
    pdf_dir = os.path.join(repo_root, "docs", "pdf")
    if not os.path.exists(pdf_dir):
        os.makedirs(pdf_dir)
        
    # 1. Detailed budget table Excel
    excel_path = os.path.join(pdf_dir, "CEF_Detailed_Budget_Table.xlsx")
    create_excel_budget(excel_path)
    
    # 2. Detailed budget table PDF
    pdf_budget_path = os.path.join(pdf_dir, "CEF_Detailed_Budget_Table.pdf")
    generate_pdf_budget_report(pdf_budget_path)
    
    # 3. Gantt chart & Timetable PDF
    pdf_gantt_path = os.path.join(pdf_dir, "CEF_Gantt_Chart_Timetable.pdf")
    generate_pdf_gantt_chart(pdf_gantt_path)
    
    # 4. Support Letter - Hellenic
    generate_letter_of_support(
        filepath=os.path.join(pdf_dir, "CEF_Letter_of_Support_Hellenic.pdf"),
        partner_name="Hellenic Submarine Telecom Authority",
        rep_name="Dimitrios Papadimitriou",
        rep_title="Director General",
        rep_address="Kifissias Avenue 99, Athens, 11524, Greece",
        partner_budget="6,000,000.00",
        match_budget="3,000,000.00",
        date="May 18, 2026"
    )
    
    # 5. Support Letter - Munich
    generate_letter_of_support(
        filepath=os.path.join(pdf_dir, "CEF_Letter_of_Support_Munich.pdf"),
        partner_name="Munich Institute of Geophysics",
        rep_name="Prof. Dr. Hans-Dieter Weber",
        rep_title="Director / Principal Geophysics Investigator",
        rep_address="Ludwig-Maximilians-Universität, Theresienstraße 41, 80333 Munich, Germany",
        partner_budget="4,500,000.00",
        match_budget="2,250,000.00",
        date="May 19, 2026"
    )
    
    # 6. Combined Support Letters
    letters_data = [
        {
            "partner_name": "Hellenic Submarine Telecom Authority",
            "rep_name": "Dimitrios Papadimitriou",
            "rep_title": "Director General",
            "rep_address": "Kifissias Avenue 99, Athens, 11524, Greece",
            "partner_budget": "6,000,000.00",
            "match_budget": "3,000,000.00",
            "date": "May 18, 2026"
        },
        {
            "partner_name": "Munich Institute of Geophysics",
            "rep_name": "Prof. Dr. Hans-Dieter Weber",
            "rep_title": "Director / Principal Geophysics Investigator",
            "rep_address": "Ludwig-Maximilians-Universität, Theresienstraße 41, 80333 Munich, Germany",
            "partner_budget": "4,500,000.00",
            "match_budget": "2,250,000.00",
            "date": "May 19, 2026"
        }
    ]
    generate_combined_letters_of_support(
        filepath=os.path.join(pdf_dir, "CEF_Letters_of_Support_Combined.pdf"),
        letters_data=letters_data
    )
    
    # 7. Ownership Control Declaration PDF
    generate_ownership_control_declaration(
        filepath=os.path.join(pdf_dir, "CEF_Ownership_Control_Declaration.pdf")
    )
    
    # 8. List of Previous Projects PDF
    generate_previous_projects_list(
        filepath=os.path.join(pdf_dir, "CEF_List_of_Previous_Projects.pdf")
    )
    
    # 9. Annual Activity Report PDF
    generate_annual_activity_report(
        filepath=os.path.join(pdf_dir, "CEF_Annual_Activity_Report.pdf")
    )
    
    # 10. Technical Specifications Annex PDF (Other Annexes)
    generate_other_annex_technical_specs(
        filepath=os.path.join(pdf_dir, "CEF_Other_Annex_Technical_Specs.pdf")
    )

def generate_annual_activity_report(filepath):
    print(f"Generating Annual Activity Report: {filepath}")
    doc = SimpleDocTemplate(
        filepath,
        pagesize=letter,
        leftMargin=54,
        rightMargin=54,
        topMargin=70,
        bottomMargin=70
    )
    
    styles = getSampleStyleSheet()
    title_style = ParagraphStyle(
        'TitleAAR', parent=styles['Heading1'],
        fontName='Helvetica-Bold', fontSize=15, leading=19,
        textColor=colors.HexColor("#1A365D"), spaceAfter=15, alignment=1
    )
    bold_body_style = ParagraphStyle(
        'BoldBodyAAR', parent=styles['BodyText'],
        fontName='Helvetica-Bold', fontSize=9.5, leading=13.5,
        textColor=colors.HexColor("#2D3748"), spaceAfter=8
    )
    body_style = ParagraphStyle(
        'BodyAAR', parent=styles['BodyText'],
        fontName='Helvetica', fontSize=9.5, leading=13.5,
        textColor=colors.HexColor("#2D3748"), spaceAfter=8
    )
    sig_style = ParagraphStyle(
        'SigAAR', parent=styles['BodyText'],
        fontName='Helvetica-Oblique', fontSize=9.5, leading=13.5,
        textColor=colors.HexColor("#4A5568")
    )
    
    story = []
    
    story.append(Paragraph("<b>ANNUAL ACTIVITY REPORT (FINANCIAL YEAR 2025)</b>", title_style))
    story.append(Paragraph("<b>AETERNA CORPORATE & OPERATIONAL REPORT</b>", ParagraphStyle('SubAAR', fontName='Helvetica-Bold', fontSize=10, leading=12, textColor=colors.HexColor("#4A5568"), alignment=1, spaceAfter=20)))
    
    story.append(Paragraph("<b>1. Executive Summary</b>", bold_body_style))
    story.append(Paragraph(
        "AETERNA is an advanced R&D technology firm specializing in zero-entropy software architectures, kernel-level eBPF security frameworks, "
        "and high-performance Edge AI systems for critical national infrastructure. Throughout 2025, AETERNA solidified its role as a key contributor "
        "to European digital sovereignty, delivering high-speed telemetry platforms and securing strategic industrial control assets.",
        body_style
    ))
    
    story.append(Paragraph("<b>2. Key Scientific & R&D Accomplishments</b>", bold_body_style))
    story.append(Paragraph(
        "• <b>Mojo Vectorized Classifier:</b> Completed the development of the O(1) vectorized neural signal separation core, achieving 35,000x execution speed-ups compared to legacy python DSP structures.<br/>"
        "• <b>eBPF Sentinel Shield:</b> Designed and deployed clean-room Linux kernel security containment loops, proving lateral SCADA isolation speeds under 1.02ms.<br/>"
        "• <b>SOP Stream Acquisition Core:</b> Built zero-latency light polarization shift data acquisition pathways, successfully tested on subsea active fiber trunks.",
        body_style
    ))
    
    story.append(Paragraph("<b>3. Active R&D Projects Portfolio</b>", bold_body_style))
    story.append(Paragraph(
        "During 2025, AETERNA actively coordinated and executed 3 major European-funded and national research initiatives:<br/>"
        "• <i>EIC Accelerator - AETERNA Core Node</i> (Grant Agreement 865986222) - €4.2M total budget.<br/>"
        "• <i>Horizon Europe Critical Infrastructure - CIRCAT-VIVI</i> - €2.8M total budget.<br/>"
        "• <i>Horizon Europe Smart Cities - AETERNA-AIGIS</i> - €5.1M total budget.",
        body_style
    ))
    
    story.append(Paragraph("<b>4. Corporate Governance & Financial Standing</b>", bold_body_style))
    story.append(Paragraph(
        "AETERNA operates under strict corporate compliance and zero foreign debt. The company's assets grew by 24% in 2025, driven by sovereign software "
        "licensing and dedicated research grants. AETERNA's operational facilities in Pomorie, Bulgaria, are certified under ISO/IEC 27001 for information "
        "security management, meeting all critical security toolbox criteria established by the EU.",
        body_style
    ))
    
    story.append(Spacer(1, 15))
    story.append(Paragraph("<b>Certified by:</b>", body_style))
    story.append(Paragraph("<b>Name:</b> Dimitar Prodromov", body_style))
    story.append(Paragraph("<b>Title:</b> Sovereign Systems Architect & Legal Rep", body_style))
    story.append(Paragraph("<b>Secure Operations Signature:</b> <code>0xCF_AAR_AETERNA_2025_APPROVED</code>", sig_style))
    
    canvas_class = make_canvas_with_metadata("Annual Activity Report", "CEF DIGITAL 2026 // AETERNA // ANNUAL REPORT 2025")
    doc.build(story, canvasmaker=canvas_class)
    print(f"Success! Annual Activity Report PDF saved at: {filepath}")

def generate_other_annex_technical_specs(filepath):
    print(f"Generating Technical Specifications Annex: {filepath}")
    doc = SimpleDocTemplate(
        filepath,
        pagesize=letter,
        leftMargin=54,
        rightMargin=54,
        topMargin=70,
        bottomMargin=70
    )
    
    styles = getSampleStyleSheet()
    title_style = ParagraphStyle(
        'TitleTS', parent=styles['Heading1'],
        fontName='Helvetica-Bold', fontSize=15, leading=19,
        textColor=colors.HexColor("#1A365D"), spaceAfter=15, alignment=1
    )
    bold_body_style = ParagraphStyle(
        'BoldBodyTS', parent=styles['BodyText'],
        fontName='Helvetica-Bold', fontSize=9.5, leading=13.5,
        textColor=colors.HexColor("#2D3748"), spaceAfter=8
    )
    body_style = ParagraphStyle(
        'BodyTS', parent=styles['BodyText'],
        fontName='Helvetica', fontSize=9.5, leading=13.5,
        textColor=colors.HexColor("#2D3748"), spaceAfter=8
    )
    
    story = []
    
    story.append(Paragraph("<b>TECHNICAL SPECIFICATIONS & HARDWARE TOPOGRAPHY ANNEX</b>", title_style))
    story.append(Paragraph("<b>AETERNA-SCW // AIGIS SUBSEA SHIELD SECURITY SYSTEM</b>", ParagraphStyle('SubTS', fontName='Helvetica-Bold', fontSize=10, leading=12, textColor=colors.HexColor("#4A5568"), alignment=1, spaceAfter=20)))
    
    story.append(Paragraph("<b>1. Subsea Optical Interrogator Hardware Specs</b>", bold_body_style))
    story.append(Paragraph(
        "The optical monitoring system (WP1) uses high-sensitivity coherent laser phase-sensing interrogator hardware. "
        "It measures polarization changes (State of Polarization - SOP) and backscatter phase shifts (Distributed Acoustic Sensing - DAS) "
        "along the active fiber trunk with the following operational constraints:<br/>"
        "• <b>Laser Coherence Length:</b> >100 km.<br/>"
        "• <b>Phase Resolution:</b> <1 microradian.<br/>"
        "• <b>Spatial Resolution:</b> 2.0 meters along a 150 km fiber segment.<br/>"
        "• <b>Sampling Frequency:</b> 10 kHz per channel, ensuring total signal integrity.",
        body_style
    ))
    
    story.append(Paragraph("<b>2. Landing-Station Edge Compute Hardware Specs</b>", bold_body_style))
    story.append(Paragraph(
        "To process the massive real-time light stream without latency, each landing-station core node operates under a bare-metal "
        "compute cluster with the following specifications:<br/>"
        "• <b>CPU Node:</b> Dual AMD EPYC 9654 (96-cores, 2.4 GHz per socket).<br/>"
        "• <b>AI Accelerator Node:</b> 4x NVIDIA H100 PCIe (80GB VRAM) for low-latency parallel vector math execution.<br/>"
        "• <b>RAM:</b> 1.5 TB DDR5 ECC Server Memory.<br/>"
        "• <b>Network Interface Card (NIC):</b> Mellanox ConnectX-7 (Dual-port 200Gb/s PCIe Gen 5.0) enabling zero-copy network DMA mapping.",
        body_style
    ))
    
    story.append(Paragraph("<b>3. Software Stack & Performance Benchmarks</b>", bold_body_style))
    story.append(Paragraph(
        "• <b>Execution Ingress Platform:</b> Written in **Zig** (`SOP_STREAM_ACQUISITION.zig`) for strict manual memory management and zero-copy register mapping.<br/>"
        "• <b>Neural Classifier core:</b> Compiled via the **Mojo** system, deploying high-performance SIMD vectorization loops to achieve a deterministic $O(1)$ signal separation execution latency of **1.14ms**.<br/>"
        "• <b>Logical Containment Module:</b> Built in **Rust** and utilizing Linux kernel-level **eBPF (Extended Berkeley Packet Filter)** Sentinel containment hooks, automating lateral SCADA traffic apoptosis in **under 1.02ms** upon acoustic tapping signature confirmation.",
        body_style
    ))
    
    story.append(Spacer(1, 15))
    story.append(Paragraph("This technical annex supports the architectural rigor and feasibility of the AETERNA-SCW infrastructure deployment.", styles['Italic']))
    
    canvas_class = make_canvas_with_metadata("Technical Annex", "CEF DIGITAL 2026 // AETERNA-SCW // TECHNICAL SPECIFICATIONS")
    doc.build(story, canvasmaker=canvas_class)
    print(f"Success! Technical Specifications Annex PDF saved at: {filepath}")

if __name__ == "__main__":
    main()
